#!/usr/bin/env python3
"""Build the dialogue workbook the writing room actually opens.

    npm run dialogue:sheet

Reads what tools/dialogue-sheet.mjs emitted and lays it out as four tabs:
PUNCH-UP (the hand-written variants, first because it is the working
document), ALL DIALOGUE (every spoken line in the game), FLAGGED (the
lines the heuristic thinks are coasting), and BY SCENE (counts).

Nothing here is authored. Fix the punch-up JSON or the manifest and rerun.
"""
import json
import pathlib

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parent.parent
DOCS = ROOT / "docs" / "dialogue"

HEAD_FILL = PatternFill("solid", fgColor="1F2933")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
SCENE_FILL = PatternFill("solid", fgColor="D9E2EC")
CURRENT_FILL = PatternFill("solid", fgColor="FFF3CD")
PICK_FILL = PatternFill("solid", fgColor="D6F5D6")
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")


def style_header(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for cell in ws[1]:
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def punchup_tab(wb, scenes):
    ws = wb.create_sheet("PUNCH-UP")
    ws.append(["Scene", "Cue id", "Character", "CURRENT LINE", "What's wrong with it",
               "House rewrite", "Tarantino", "McDonagh", "Houser (GTA)", "Coen",
               "PICK", "Your notes"])
    for scene in scenes:
        row = ws.max_row + 1
        ws.cell(row=row, column=1, value=f"{scene['scene']}  —  {scene['reference']}").font = Font(bold=True, size=12)
        for c in range(1, 13):
            ws.cell(row=row, column=c).fill = SCENE_FILL
        ws.row_dimensions[row].height = 18

        row = ws.max_row + 1
        ws.cell(row=row, column=1, value="Diagnosis").font = Font(bold=True, italic=True)
        d = ws.cell(row=row, column=2, value=scene["diagnosis"])
        d.alignment = WRAP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
        ws.row_dimensions[row].height = 90

        for line in scene.get("lines", []):
            ws.append([scene["scene"], line["cue"], line.get("character", ""),
                       line.get("current", ""), line.get("why", ""), line.get("house", ""),
                       line.get("tarantino", ""), line.get("mcdonagh", ""),
                       line.get("houser", ""), line.get("coen", ""), "", ""])
            r = ws.max_row
            for c in range(1, 13):
                ws.cell(row=r, column=c).alignment = WRAP
            ws.cell(row=r, column=4).fill = CURRENT_FILL
            ws.cell(row=r, column=11).fill = PICK_FILL

        if scene.get("newMaterial"):
            nm = scene["newMaterial"]
            row = ws.max_row + 2
            ws.cell(row=row, column=1, value="NEW MATERIAL").font = Font(bold=True, size=12)
            ws.cell(row=row, column=2, value=nm.get("note", "")).alignment = WRAP
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
            ws.row_dimensions[row].height = 60
            for pitch in nm.get("pitches", []):
                ws.append(["", "", "", "", "", pitch, "", "", "", "", "", ""])
                ws.cell(row=ws.max_row, column=6).alignment = WRAP

    style_header(ws, [22, 34, 14, 46, 46, 52, 52, 52, 52, 52, 12, 28])
    return ws


def all_tab(wb, rows):
    ws = wb.create_sheet("ALL DIALOGUE")
    ws.append(["Row", "Scene", "Character", "Cue id", "Filename", "Recorded", "Direction",
               "CURRENT LINE", "Words", "Flags", "House rewrite", "Tarantino", "McDonagh",
               "Houser (GTA)", "Coen", "PICK", "Notes"])
    for r in rows:
        ws.append([r["row"], r["scene"], r["character"], r["cue"], r["file"], r["recorded"],
                   r["direction"], r["current"], r["words"], r["flags"], r["punchUp"],
                   r["tarantino"], r["mcdonagh"], r["houser"], r["coen"], r["pick"], r["notes"]])
        i = ws.max_row
        ws.cell(row=i, column=8).alignment = WRAP
        for c in (11, 12, 13, 14, 15, 17):
            ws.cell(row=i, column=c).alignment = WRAP
        ws.cell(row=i, column=8).fill = CURRENT_FILL
        ws.cell(row=i, column=16).fill = PICK_FILL
    style_header(ws, [6, 22, 16, 38, 40, 10, 30, 60, 7, 16, 48, 48, 48, 48, 48, 10, 40])
    return ws


def flagged_tab(wb, rows):
    ws = wb.create_sheet("FLAGGED")
    ws.append(["Row", "Scene", "Character", "Cue id", "CURRENT LINE", "Words", "Flags",
               "Punched up?"])
    for r in rows:
        if not r["flags"]:
            continue
        ws.append([r["row"], r["scene"], r["character"], r["cue"], r["current"], r["words"],
                   r["flags"], "yes" if r["punchUp"] else ""])
        ws.cell(row=ws.max_row, column=5).alignment = WRAP
    style_header(ws, [6, 22, 16, 38, 70, 7, 20, 12])
    return ws


def scene_tab(wb, rows):
    ws = wb.create_sheet("BY SCENE")
    ws.append(["Scene", "Lines", "Flagged", "Punched up", "Recorded", "Unrecorded"])
    scenes = {}
    for r in rows:
        s = scenes.setdefault(r["scene"], [0, 0, 0, 0, 0])
        s[0] += 1
        s[1] += 1 if r["flags"] else 0
        s[2] += 1 if r["punchUp"] else 0
        s[3] += 1 if r["recorded"] == "yes" else 0
        s[4] += 1 if r["recorded"] == "no" else 0
    for name, s in sorted(scenes.items(), key=lambda kv: -kv[1][0]):
        ws.append([name, s[0], s[1], s[2], s[3], s[4]])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = TOP
    style_header(ws, [30, 8, 10, 12, 11, 12])
    return ws


def main():
    rows = json.loads((DOCS / "DIALOGUE-MASTER.json").read_text())
    scenes = json.loads((DOCS / "DIALOGUE-PUNCHUP.json").read_text())

    wb = Workbook()
    wb.remove(wb.active)
    punchup_tab(wb, scenes)
    all_tab(wb, rows)
    flagged_tab(wb, rows)
    scene_tab(wb, rows)

    out = DOCS / "SQUATCH-SMASH-DIALOGUE.xlsx"
    wb.save(out)
    punched = sum(1 for r in rows if r["punchUp"])
    print(f"{out.relative_to(ROOT)} — {len(rows)} lines, {punched} with punch-up variants, "
          f"{len(scenes)} scene write-ups.")


if __name__ == "__main__":
    main()
