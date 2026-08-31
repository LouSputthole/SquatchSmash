#!/usr/bin/env python3
"""Build the dialogue + audio pass workbook the owner asked for.

    npm run dialogue:pass && python3 tools/dialogue-workbook.py

Reads docs/dialogue/DIALOGUE-PASS.json (every spoken line, in spoken order,
by beat) and the generated radio audit, and writes

    docs/SQUATCHSMASH-DIALOGUE-AND-AUDIO-PASS.xlsx

The findings on LEVEL 1 and LEVEL 2 are hand-written, not inferred: each one
names the evidence that supports it. The REWRITE column is the second version
of a line the owner picks from -- same character, same beat, better writing.
"""
import json
import os
import zipfile
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PASS_JSON = os.path.join(ROOT, 'docs/dialogue/DIALOGUE-PASS.json')
RADIO_XLSX = os.path.join(ROOT, 'docs/audits/SQUATCHSMASH-RADIO-AUDIT.xlsx')
OUT = os.path.join(ROOT, 'docs/SQUATCHSMASH-DIALOGUE-AND-AUDIO-PASS.xlsx')

FONT = 'Arial'
INK = '1F2430'
RULE = 'B7BDC8'
HEAD_FILL = '1F2430'
L1_FILL = 'F6D2CE'
L2_FILL = 'FDE8C8'
OK_FILL = 'DCE9DA'
BAND = 'F2F4F7'

thin = Side(style='thin', color=RULE)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def head(ws, row, labels, widths):
    for i, (label, width) in enumerate(zip(labels, widths), start=1):
        c = ws.cell(row=row, column=i, value=label)
        c.font = Font(name=FONT, bold=True, color='FFFFFF', size=10)
        c.fill = PatternFill('solid', fgColor=HEAD_FILL)
        c.alignment = Alignment(vertical='center', wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[row].height = 30


def title(ws, text, sub, span):
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(name=FONT, bold=True, size=15, color=INK)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    s = ws.cell(row=2, column=1, value=sub)
    s.font = Font(name=FONT, size=9.5, italic=True, color='5A6472')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    ws.row_dimensions[1].height = 21


def body(ws, row, values, fill=None, wrap_cols=()):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=9.5, color=INK)
        c.border = BORDER
        c.alignment = Alignment(
            vertical='top', wrap_text=(i in wrap_cols))
        if fill:
            c.fill = PatternFill('solid', fgColor=fill)


rows = json.load(open(PASS_JSON, encoding='utf8'))

# ---------------------------------------------------------------- findings --
# Cue -> (level, why, rewrite). Every entry was checked against the source
# named in the WHY column; nothing here is guessed from the line alone.
FLAGS = {
    'vo.door.refusal.golf_call.1': (
        'L1',
        'DEAD LINE. Beat 12 is now NEW_SPACE_LOU_CALL and Lou rings the night '
        'before, so the door never reaches this key. src/core/apartment-story.js:74 '
        'says so outright: kept only because the take is on disk and the orphan '
        'gate counts unclaimed files.',
        'RETIRE the take. If you want the beat covered, the honest replacement is '
        'already there: new_space_call.'),
    'vo.door.refusal.golf_return.1': (
        'L1',
        'DEAD LINE, and it contradicts the Home Ladder. Coming home from the round '
        'does not happen — the round ends at the new address (beat 14), and he never '
        'goes back down a rung. src/core/apartment-story.js:74.',
        'RETIRE. Nothing replaces it: there is no "home from golf" in the route.'),
    'vo.door.refusal.date_call.1': (
        'L1',
        'Retired with the duplicate apartment call; src/core/apartment-story.js '
        'keeps it only as "a claimed legacy recording until the audio-debt pass '
        'removes the physical take."',
        'RETIRE with the audio-debt pass.'),
    'vo.door.refusal.sleep_after_date.1': (
        'L1',
        'Same legacy retention as date_call. The date now ends at the luxury '
        'apartment with Margo, not in the starter flat.',
        'RETIRE, or repoint to the stayover if you want a beat-16 bedtime line.'),
    'vo.silentsquatch.return.briefing.lou.instrument': (
        'L1',
        'RIGHT LINE, WRONG NAMESPACE. This is Lou\'s wrong-city reveal and it plays '
        'at the REPAIRED mansion (beat 25), but it lives under vo.silentsquatch.* '
        '(beat 22). Any prefix-to-scene tool files beat 25 as having no dialogue — '
        'this one did until the route was corrected.',
        'Keep the line word for word. Rename the prefix to vo.mansionreturn.* so the '
        'beat owns its own dialogue.'),
    'vo.bing.full.james_blond.qCase.line.pcyjzg': (
        'L1?',
        'OWNER QUESTION, not a defect. The James Blond bit at Bada Bing I (beat 2) '
        'names "the silver case" — the same object as the beat-20 MacGuffin, '
        'eighteen beats early. Verified as a separate gag, but a player cannot know '
        'that.',
        'If it is a deliberate plant, leave it. If not, change the noun: '
        '"The briefcase went where briefcases go."'),
}

# Level 2 is a shape, not a cue list: these are the exemplars for each pattern.
L2_ROWS = [
    ('Flat interjection carrying a take',
     '416 lines are two words or fewer. "Right." x7, "No." x7, "Okay." x5, '
     '"Yeah." x4 — each one is a separate voice take doing almost no work.',
     'Where the line is a beat of silence, cut the take and let the room play. '
     'Where it is a reaction, give it a noun: "Right." -> "Right. Course it is."',
     'Beats 1, 13, 22, 28, 29'),
    ('Functional / HUD-speak opener',
     '94 lines open with an instruction verb or a filler acknowledgement — '
     '"Okay. Okay, we\'re doing this.", "Right, let\'s have a look." They read as '
     'stage directions the player can hear.',
     'Say the thing he is actually looking at. "Right, let\'s have a look." -> '
     '"Whatever\'s in there, it\'s been in there a while."',
     'Beats 0, 1'),
    ('Same words, many takes',
     '73 exact-duplicate line texts across 213 cues. "Another time." appears 18 '
     'times in Bada Bing I alone.',
     'Keep one take, vary the rest. A refusal bank that says the same four words '
     'eighteen times is the one place the Family stops sounding serious.',
     'Beat 2'),
    ('VERIFIED INTENTIONAL — do not "fix"',
     'Shubenator\'s "Hey guys, what\'s going on?" recurs 7 times across beats 2, 8, '
     '11.5, 22 and 24. The cue names are signature.cheerful / signature.gleeful — '
     'it is a running gag, and it was nearly filed as a placeholder here.',
     'LEAVE IT. Noted so the next pass does not delete the joke.',
     'Beats 2, 8, 11.5, 22, 24'),
]

SILENT_BEATS = [
    ('11', 'Return to Old Apartment',
     'Normal life, and it does not feel the same — with nothing said about it.'),
    ('14', 'Luxury Apartment',
     'Lou\'s gift. THE STARTER FLAT GOES DARK HERE. The biggest promotion in the '
     'game, and the only three spoken cues the luxury apartment owns are door '
     'refusals.'),
    ('17', 'Luxury Apartment Morning',
     'She leaves, then the phone. The morning after is silent.'),
    ('19', 'Luxury Apartment Return',
     'Quiet, then a call about something sensitive. The quiet is literal.'),
]

MUSIC = [
    ('1', '3 -> 4', 'The drive out of town after the Squatchfather',
     'The driver takes him OUT OF TOWN after his first kill. Night county road, '
     '2 hours 20 minutes of it, and no score at all.',
     'Same shape as "Driving THE TAKE" and "Driving to the Jerky Hotel" — the two '
     'drives already scored. This is the drive that earns one most: it is the first '
     'one where he has done something he cannot take back.',
     'Low, unhurried, mostly bass and brushed drums. No melody with an opinion. It '
     'should sound like the car, not like a verdict.', 'HIGHEST'),
    ('2', '28', 'Seff\'s forty-two-minute ride',
     'The last ride before he is made, with something in the trunk. src/specialmeeting '
     'wires no music at all.',
     'It is the longest continuous drive in the campaign and the only one with a '
     'body in it. Also the last scene before the credits beat.',
     'Start almost ambient and let one instrument arrive around the halfway mark. The '
     'trunk is the joke and the dread; the music should know only the dread.', 'HIGHEST'),
    ('3', '14', 'Walking into the luxury apartment for the first time',
     'Lou\'s gift, and the starter flat goes dark forever. Currently one of four '
     'beats in the game with no dialogue AND no music.',
     'The reward for THE TAKE. Right now the promotion plays in silence, which reads '
     'as a missing scene rather than a quiet one. Music can carry the whole beat '
     'without a single line being written.',
     'Warm, wide, slightly too grand for the room — the score believes he has made it '
     'more than he does.', 'HIGH'),
    ('4', '7', 'The pyre and the blackout',
     'End of Act One: executions, the pyre, nightfall at 20:45, then the lights go.',
     'The act break. The cabin wires no music of any kind today.',
     'One sustained low tone that arrives with the fire and does not resolve. Let the '
     'blackout cut it dead rather than fading it.', 'HIGH'),
    ('5', '9', 'The burial',
     'Billy goes in the ground and he is told not to go home tonight.',
     'The graveyard wires no music. It is the shortest scene in the chapter (22 lines) '
     'and the most exposed by silence.',
     'Something almost liturgical played badly — a small ensemble that has done this '
     'before and is not moved any more.', 'MEDIUM'),
    ('6', '29', 'Made, and the credits roll',
     'The ceremony lands at 19:00 and the game ends.',
     'The credits are the last thing anyone hears. src/initiation wires no music.',
     'The one place to state a real theme. Whatever you have been withholding all '
     'game, play it here with the full arrangement.', 'MEDIUM'),
]

wb = Workbook()

# ----------------------------------------------------------------- README ---
ws = wb.active
ws.title = 'README'
title(ws, 'SQUATCHSMASH — DIALOGUE & AUDIO PASS',
      'Generated from the live source. Regenerate: npm run dialogue:pass && '
      'python3 tools/dialogue-workbook.py', 2)
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 108
guide = [
    ('WHAT IS IN HERE', ''),
    ('DIALOGUE — ALL', f'All {len(rows):,} spoken lines, in the order they are spoken, '
     'grouped by the 31-beat spine and named by character. Filter the Beat column to '
     'read one scene at a time.'),
    ('BEAT INDEX', 'Every beat with its line count. Four beats have none — see LEVEL 1.'),
    ('LEVEL 1 — FIX FIRST', 'Lines that are stale, unreachable, or contradict the story '
     'as it stands now. Each row names the source that proves it.'),
    ('LEVEL 2 — WEAK', 'Lines that work but are not pulling their weight, grouped by '
     'the pattern they share rather than listed one by one.'),
    ('RADIO BY SCENE', 'What plays in every scene, and how it behaves — trigger, stop, '
     'volume, ducking, overlap risk.'),
    ('MUSIC PLACEMENTS', 'Six places to put a background song, ranked, with a brief for '
     'each.'),
    ('', ''),
    ('HOW TO USE THE REWRITE COLUMN',
     'Every flagged line carries a second version. Pick the one you want and say so — '
     'the line lives in the manifest, so changing it is one edit plus a regenerate.'),
    ('', ''),
    ('WHAT IS AUTHORITATIVE', 'docs/CAMPAIGN-STORY-BIBLE.md, then src/core/campaign-spine.js. '
     'Where this workbook and the bible disagree, the bible is right.'),
    ('ORDERING', 'Each vo:<scene> generator walks its script in authored order and appends '
     'rows as it goes, so a cue\'s position in assets/sfx/manifest.json IS its spoken '
     'order within its scene. The ORDERING column says which scenes are a straight run '
     'and which are a tree or a bark bank.'),
    ('COVERAGE, HONESTLY', 'The 4,229 lines were machine-screened in full for staleness '
     'against every settled story rule, for duplicates, for HUD-speak and for length. '
     'LEVEL 1 is hand-verified line by line against the source. LEVEL 2 names the '
     'patterns and their exemplars, not all 583 matching lines.'),
]
r = 4
for k, v in guide:
    a = ws.cell(row=r, column=1, value=k)
    a.font = Font(name=FONT, bold=bool(k), size=10, color=INK)
    a.alignment = Alignment(vertical='top', wrap_text=True)
    b = ws.cell(row=r, column=2, value=v)
    b.font = Font(name=FONT, size=10, color=INK)
    b.alignment = Alignment(vertical='top', wrap_text=True)
    if v and len(v) > 90:
        ws.row_dimensions[r].height = 30
    r += 1

# ------------------------------------------------------------ BEAT INDEX ---
ws = wb.create_sheet('BEAT INDEX')
title(ws, 'BEAT INDEX', 'Line count per beat. A beat with no lines is a beat that plays '
      'in silence.', 4)
head(ws, 4, ['Beat', 'Chapter', 'Scene', 'Spoken lines'], [10, 26, 40, 14])
counts = {}
order_seen = []
for row_ in rows:
    key = (row_['beat'], row_['chapter'], row_['beatTitle'])
    if key not in counts:
        counts[key] = 0
        order_seen.append(key)
    counts[key] += 1
silent_ids = {b for b, _, _ in SILENT_BEATS}
r = 5
for key in order_seen:
    beat, chapter, scene = key
    body(ws, r, [beat, chapter, scene, counts[key]], wrap_cols=(3,))
    r += 1
for beat, scene, _why in SILENT_BEATS:
    body(ws, r, [beat, '(see LEVEL 1)', scene, 0], fill=L1_FILL, wrap_cols=(3,))
    r += 1
ws.freeze_panes = 'A5'

# --------------------------------------------------------- DIALOGUE — ALL ---
ws = wb.create_sheet('DIALOGUE — ALL')
title(ws, 'EVERY SPOKEN LINE, IN ORDER',
      f'{len(rows):,} lines. Filter Beat to read one scene. FLAG marks a line carried '
      'on LEVEL 1 or LEVEL 2.', 11)
cols = ['Beat', 'Chapter', 'Scene', '#', 'Character', 'Voice', 'Line', 'Words',
        'Ordering', 'Flag', 'Rewrite / note']
head(ws, 4, cols, [8, 22, 26, 6, 30, 12, 78, 7, 30, 7, 60])
r = 5
for row_ in rows:
    flag = FLAGS.get(row_['cue'])
    body(ws, r, [
        row_['beat'], row_['chapter'], row_['beatTitle'], row_['order'],
        row_['character'], row_['voice'], row_['line'], row_['words'],
        row_['ordering'], flag[0] if flag else '',
        flag[2] if flag else '',
    ], fill=L1_FILL if flag else None, wrap_cols=(7, 11))
    r += 1
ws.freeze_panes = 'E5'
ws.auto_filter.ref = f'A4:K{r - 1}'

# -------------------------------------------------------------- LEVEL 1 ----
ws = wb.create_sheet('LEVEL 1 — FIX FIRST')
title(ws, 'LEVEL 1 — STALE, DEAD, OR CONTRADICTED BY THE STORY',
      'Hand-verified. The WHY column names the file that proves it.', 6)
head(ws, 4, ['#', 'Beat', 'Where', 'The line as it stands', 'Why it is a problem',
             'Rewrite / decision'], [5, 9, 34, 52, 62, 58])
r = 5
n = 1
for beat, scene, why in SILENT_BEATS:
    body(ws, r, [n, beat, scene, '(nothing — this beat has no dialogue at all)', why,
                 'WRITE THE BEAT. See the note below for beat 14.'],
         fill=L1_FILL, wrap_cols=(3, 4, 5, 6))
    ws.row_dimensions[r].height = 46
    r += 1
    n += 1
by_cue = {row_['cue']: row_ for row_ in rows}
for cue, (level, why, rewrite) in FLAGS.items():
    row_ = by_cue.get(cue)
    if not row_ or level not in ('L1', 'L1?'):
        continue
    body(ws, r, [n, row_['beat'], cue, row_['line'], why, rewrite],
         fill=L1_FILL if level == 'L1' else L2_FILL, wrap_cols=(3, 4, 5, 6))
    ws.row_dimensions[r].height = 46
    r += 1
    n += 1
extra = [
    ('21', 'vo.silvercase.car.* / .arrival.*',
     '(12 cues)',
     'These play on the ride TO the mansion (beat 21) but sit in the beat-20 '
     'namespace. Beat 21 read as silent until the route was corrected.',
     'Split the prefix, or accept it and keep the mapping in tools/dialogue-pass.mjs.'),
    ('27', 'vo.specialmeeting.tony.idle_before.*',
     '(21 cues)',
     'Him alone in the luxury flat waiting for the phone — that is beat 27, not the '
     'beat-28 ride the prefix implies.',
     'Rename to vo.specialmeeting.apartment.* so the two halves of the beat are '
     'separable.'),
    ('7', 'vo.call.booski.cabin_billy.*',
     '(cues)',
     'Booski rings the cabin twice. The Captain call is beat 5; this summons is the '
     'END of beat 7, on Day 4 after the blackout. One prefix, eleven hours apart.',
     'Name the second one for its day, not its caller.'),
    ('—', 'npm run check:reachability',
     '(gate coverage)',
     'The reachability gate covers 5 scenes of 30 — mansion, initiation, beefrun, '
     'enolasquatch, bing. The apartment door-refusal bank, where all four confirmed '
     'dead lines live, is not gated by anything.',
     'Extend the gate to the apartment bank first; it is where the dead lines '
     'actually are.'),
]
for beat, where, line, why, rewrite in extra:
    body(ws, r, [n, beat, where, line, why, rewrite], fill=L1_FILL,
         wrap_cols=(3, 4, 5, 6))
    ws.row_dimensions[r].height = 46
    r += 1
    n += 1
ws.freeze_panes = 'A5'

# -------------------------------------------------------------- LEVEL 2 ----
ws = wb.create_sheet('LEVEL 2 — WEAK')
title(ws, 'LEVEL 2 — LINES THAT ARE NOT PULLING THEIR WEIGHT',
      'Grouped by pattern. The last row is a line that LOOKS like a defect and is not.', 4)
head(ws, 4, ['Pattern', 'What the screen found', 'Rewrite approach', 'Where'],
     [34, 66, 66, 26])
r = 5
for pattern, found, approach, where in L2_ROWS:
    fill = OK_FILL if pattern.startswith('VERIFIED') else L2_FILL
    body(ws, r, [pattern, found, approach, where], fill=fill, wrap_cols=(1, 2, 3, 4))
    ws.row_dimensions[r].height = 58
    r += 1
ws.freeze_panes = 'A5'

# -------------------------------------------------------- RADIO BY SCENE ---
ws = wb.create_sheet('RADIO BY SCENE')
title(ws, 'RADIO & MUSIC, SCENE BY SCENE',
      'From tools/radio-audit.mjs — what plays, when it starts and stops, and how it '
      'behaves against dialogue.', 9)
radio_cols = ['Beat', 'Location or venue', 'Source', 'Station', 'Trigger',
              'Stop condition', 'Volume', 'Ducking behavior', 'Overlap risk']
head(ws, 4, radio_cols, [26, 30, 30, 20, 40, 30, 24, 40, 28])
r = 5
if os.path.exists(RADIO_XLSX):
    z = zipfile.ZipFile(RADIO_XLSX)
    ns = {'m': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
    shared = []
    root = ET.fromstring(z.read('xl/sharedStrings.xml'))
    for si in root.findall('m:si', ns):
        shared.append(''.join(
            t.text or '' for t in si.iter(
                '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t')))
    sheet = ET.fromstring(z.read('xl/worksheets/sheet1.xml'))
    grid = []
    for row_el in sheet.iter(
            '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row'):
        cells = []
        for c in row_el:
            v = c.find('m:v', ns)
            cells.append('' if v is None else (
                shared[int(v.text)] if c.get('t') == 's' else v.text))
        grid.append(cells)
    header = grid[3]
    idx = {name: header.index(name) for name in [
        'Beat', 'Location or venue', 'Radio or music source', 'Station', 'Trigger',
        'Stop condition', 'Volume', 'Ducking behavior', 'Overlap risk'] if name in header}
    seen = set()
    for data in grid[4:]:
        def get(name):
            i = idx.get(name)
            return data[i] if i is not None and i < len(data) else ''
        key = (get('Beat'), get('Radio or music source'), get('Station'))
        if key in seen:
            continue
        seen.add(key)
        body(ws, r, [get('Beat'), get('Location or venue'),
                     get('Radio or music source'), get('Station'), get('Trigger'),
                     get('Stop condition'), get('Volume'), get('Ducking behavior'),
                     get('Overlap risk')], wrap_cols=(2, 3, 5, 6, 7, 8, 9))
        ws.row_dimensions[r].height = 32
        r += 1
ws.freeze_panes = 'A5'

# ----------------------------------------------------- MUSIC PLACEMENTS ----
ws = wb.create_sheet('MUSIC PLACEMENTS')
title(ws, 'SIX PLACES TO PUT A BACKGROUND SONG',
      'Ranked. Every one of these beats currently wires no music of any kind — '
      'verified against src/.', 7)
head(ws, 4, ['Rank', 'Beat', 'The moment', 'What is there now', 'Why here',
             'Brief for the track', 'Priority'],
     [6, 10, 38, 52, 60, 62, 12])
r = 5
for rank, beat, moment, now, why, brief, prio in MUSIC:
    body(ws, r, [rank, beat, moment, now, why, brief, prio],
         fill=OK_FILL if prio == 'HIGHEST' else None, wrap_cols=(3, 4, 5, 6))
    ws.row_dimensions[r].height = 58
    r += 1
r += 1
note = ws.cell(row=r, column=1, value='ALREADY SCORED (do not duplicate)')
note.font = Font(name=FONT, bold=True, size=10, color=INK)
r += 1
for text in [
    'Beat 6 Beef Run — "Can\'t You Hear Me Knocking" (src/beefrun/mission.js)',
    'Beat 10 Jerky Motel — "Driving to the Jerky Hotel" (src/motel/audio.js)',
    'Beat 11.5 THE TAKE — "Driving THE TAKE" + "Codename: Sasquatch" (src/heist/music.js)',
    'Beat 15 Front & Center — "Big Feet on the Dance Floor" + "Coco Cabana" (src/silver)',
    'Beat 24 SQUATCHOLA GAY — pre-drop approach, escape after drop, "Fortunate Son"',
    'Character cues — "Sensi Lou" on Lou\'s office door, "Baby Snakes" on Booski '
    '(src/core/signature-music.js). Neither recording is on disk yet; both fall back.',
]:
    c = ws.cell(row=r, column=1, value=text)
    c.font = Font(name=FONT, size=9.5, color='5A6472')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1
ws.freeze_panes = 'A5'

wb.save(OUT)
print(f'wrote {OUT}')
print(f'  {len(rows):,} dialogue rows · {len(FLAGS)} flagged cues · '
      f'{len(SILENT_BEATS)} silent beats · {len(MUSIC)} music placements')
